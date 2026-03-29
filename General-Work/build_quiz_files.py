import random
import csv
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

random.seed(42)  # reproducible shuffle for baked-in option order

OUT = '/home/claude/General-Work'
os.makedirs(f'{OUT}/_templates', exist_ok=True)
os.makedirs(f'{OUT}/Dairy', exist_ok=True)
os.makedirs(f'{OUT}/Landscaping', exist_ok=True)

# ── COLORS ──────────────────────────────────────────────────────────────────
HDR_FILL   = PatternFill('solid', start_color='1A5C1A')   # dark green
HDR_FONT   = Font(name='Arial', bold=True, color='FFFFFF', size=10)
NOTE_FILL  = PatternFill('solid', start_color='FFF8E1')   # pale yellow
NOTE_FONT  = Font(name='Arial', italic=True, color='555555', size=9)
ID_FILL    = PatternFill('solid', start_color='E8F5E9')   # pale green
BODY_FONT  = Font(name='Arial', size=10)
ID_FONT    = Font(name='Arial', size=10, color='1A5C1A')
WRAP       = Alignment(wrap_text=True, vertical='top')
THIN       = Side(style='thin', color='A5D6A7')
BORDER     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Column definitions: (header_label, width, note)
COLUMNS = [
    ('quiz_id',                  12, 'e.g. DAIRY-001'),
    ('quiz_title',               22, 'Display title shown in <h1>'),
    ('scenario_id (optional)',   18, 'e.g. DAIRY-001-S01 — leave blank if no scenarios'),
    ('scenario_title (optional)',22, 'Short label for the scenario selector button'),
    ('scenario_narrative_en (optional)', 40, 'English narrative shown above questions'),
    ('scenario_narrative_es (optional)', 40, 'Spanish narrative (shown below English, in <em>)'),
    ('question_id',              16, 'e.g. DAIRY-001-Q01'),
    ('sentence_en',              45, 'Full English sentence. Wrap target word in [brackets].'),
    ('keyword',                  18, 'Plain-text keyword shown in feedback (Format A only; leave blank for Format B)'),
    ('correct',                  40, 'Correct answer text'),
    ('option_b',                 40, 'Distractor B'),
    ('option_c',                 40, 'Distractor C'),
    ('option_d',                 40, 'Distractor D (leave blank for 3-option questions)'),
    ('tts_lang',                 10, 'e.g. en-US (default) or es-US'),
    ('notes',                    30, 'Author notes — not exported to HTML or CSV library'),
]

def style_sheet(ws, col_defs):
    # Row 1: column headers
    for c, (label, width, _) in enumerate(col_defs, 1):
        cell = ws.cell(row=1, column=c, value=label)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.row_dimensions[1].height = 30

    # Row 2: notes (pale yellow)
    for c, (_, _, note) in enumerate(col_defs, 1):
        cell = ws.cell(row=2, column=c, value=note)
        cell.font = NOTE_FONT
        cell.fill = NOTE_FILL
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        cell.border = BORDER
    ws.row_dimensions[2].height = 28

    ws.freeze_panes = 'A3'

def write_data_row(ws, row_num, values, id_cols=(0, 2, 6)):
    for c, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=c, value=val if val is not None else '')
        cell.font = ID_FONT if (c-1) in id_cols else BODY_FONT
        cell.fill = ID_FILL if (c-1) in id_cols else PatternFill()
        cell.alignment = WRAP
        cell.border = BORDER

def shuffle_options(correct, distractors):
    """Shuffle correct + distractors once; return (ordered_options, correct_index)."""
    opts = [correct] + distractors
    random.shuffle(opts)
    return opts, opts.index(correct)

# ── TEMPLATE ────────────────────────────────────────────────────────────────
def build_template():
    wb = Workbook()
    ws = wb.active
    ws.title = 'quiz_data'
    style_sheet(ws, COLUMNS)
    # One example row to show format
    example = [
        'TOPIC-001', 'My Quiz Title',
        'TOPIC-001-S01', 'Scenario Title',
        'English narrative for this scenario.',
        'Narrative in Spanish.',
        'TOPIC-001-Q01',
        'Use the [shovel] to dig the hole.',
        'shovel',
        'la pala', 'el rastrillo', 'la manguera', 'la azada',
        'en-US', 'Example row — delete before use'
    ]
    write_data_row(ws, 3, example)
    ws.row_dimensions[3].height = 20
    wb.save(f'{OUT}/_templates/quiz-template.xlsx')
    print('Template saved.')

# ── DAIRY SOURCE DATA ────────────────────────────────────────────────────────
# Raw data: (scenario_id, scenario_title, narrative_en, narrative_es, questions[])
# Each question: (q_id, sentence_en, keyword, correct, [distractors])

DAIRY_SCENARIOS = [
  ('DAIRY-001-S01','Clean the Barn',
   'Milking is done. Now you need to run the alley scraper, wash the aisles, connect the hose, and take out the trash before the next shift.',
   'El ordeño ha terminado. Ahora tiene que poner en marcha el raspador de pasillos, lavar los pasillos, conectar la manguera y tirar la basura antes del siguiente turno.',
   [
    ('DAIRY-001-Q01','[Run] the alley scraper to clean manure from the aisles.','run','poner en marcha',['correr','subir']),
    ('DAIRY-001-Q02','Run the [alley scraper] to clean manure from the aisles.','alley scraper','raspador de pasillos',['empacadora','cargadora']),
    ('DAIRY-001-Q03','Run the alley scraper to [clean] manure from the aisles.','clean','limpiar',['lavar','arreglar']),
    ('DAIRY-001-Q04','Run the alley scraper to clean [manure] from the aisles.','manure','estiércol',['heno','arena']),
    ('DAIRY-001-Q05','Run the alley scraper to clean manure from the [aisles].','aisles','pasillos',['corrales','establos']),
    ('DAIRY-001-Q06','[Connect] the hose to the faucet and the nozzle to the hose.','connect','conectar',['fijar','encender']),
    ('DAIRY-001-Q07','Connect the [hose] to the faucet and the nozzle to the hose.','hose','manguera',['cubeta','boquilla']),
    ('DAIRY-001-Q08','Connect the hose to the [faucet] and the nozzle to the hose.','faucet','grifo',['pasillo','basurero']),
    ('DAIRY-001-Q09','Connect the hose to the faucet and the [nozzle] to the hose.','nozzle','boquilla',['manguera','cuerda']),
    ('DAIRY-001-Q10','[Wash down] the aisle after milking.','wash down','lavar',['limpiar','rociar']),
    ('DAIRY-001-Q11','Wash down the [aisle] after milking.','aisle','pasillo',['corral','establo']),
    ('DAIRY-001-Q12','Wash down the aisle after [milking].','milking','el ordeño',['la cosecha','el turno']),
    ('DAIRY-001-Q13','[Throw out] the trash — put the trash bags in the trashcan.','throw out','tirar',['limpiar','poner']),
    ('DAIRY-001-Q14','Throw out the trash — put the [trash bags] in the trashcan.','trash bags','bolsas de basura',['cubetas','mangueras']),
    ('DAIRY-001-Q15','Throw out the trash — put the trash bags in the [trashcan].','trashcan','basurero',['corral','establo']),
    ('DAIRY-001-Q16','Use the [bucket] to clean up the milk spill in the aisle.','bucket','cubeta',['pala','manguera']),
    ('DAIRY-001-Q17','Use the bucket to [clean up] the milk spill in the aisle.','clean up','limpiar',['lavar','raspar']),
    ('DAIRY-001-Q18','Use the bucket to clean up the [milk] spill in the aisle.','milk','leche',['agua','estiércol']),
   ]),
  ('DAIRY-001-S02','Feed the Herd',
   "It's feeding time. You need to start the mixer, prepare the ration, load hay bales, fill the troughs, and make sure the cows have water.",
   'Es hora de alimentar. Tiene que encender la mezcladora, preparar la ración, cargar las pacas de heno, llenar los comederos y asegurarse de que las vacas tengan agua.',
   [
    ('DAIRY-001-Q19','[Start] the mixer to prepare the total mixed ration.','start','encender',['poner en marcha','conectar']),
    ('DAIRY-001-Q20','Start the [mixer] to prepare the total mixed ration.','mixer','mezcladora',['cargadora','picadora']),
    ('DAIRY-001-Q21','[Cut] the string on the bale before putting it in the mixer.','cut','cortar',['doblar','apretar']),
    ('DAIRY-001-Q22','Cut the [string] on the bale before putting it in the mixer.','string','cuerda',['alambre','cadena']),
    ('DAIRY-001-Q23','Cut the string on the [bale] before putting it in the mixer.','bale','paca',['comedero','remolque']),
    ('DAIRY-001-Q24','Cut the string on the bale before [putting] it in the mixer.','putting','poner',['lavar','arreglar']),
    ('DAIRY-001-Q25','Put the [feed] in the trough before the cows come in.','feed','alimento',['agua','ensilaje']),
    ('DAIRY-001-Q26','Put the feed in the [trough] before the cows come in.','trough','comedero',['pasillo','establo']),
    ('DAIRY-001-Q27','[Give] the cows water in the troughs.','give','dar',['poner','subir']),
    ('DAIRY-001-Q28','Give the cows [water] in the troughs.','water','agua',['leche','alimento']),
    ('DAIRY-001-Q29','Use the [truck] to haul the feed to the far barn.','truck','camión de carga',['tractor','remolque']),
    ('DAIRY-001-Q30','Use the truck to [haul] the feed to the far barn.','haul','transportar',['limpiar','poner']),
    ('DAIRY-001-Q31','Use the truck to haul the feed to the far [barn].','barn','establo',['corral','laguna']),
    ('DAIRY-001-Q32','[Load] the hay bales onto the wagon before the rain.','load','cargar',['sacar','mover']),
    ('DAIRY-001-Q33','Load the [hay] bales onto the wagon before the rain.','hay','heno',['maíz','ensilaje']),
    ('DAIRY-001-Q34','Load the hay [bales] onto the wagon before the rain.','bales','pacas',['cubetas','llantas']),
    ('DAIRY-001-Q35','Load the hay bales onto the [wagon] before the rain.','wagon','remolque',['cargadora','tractor']),
   ]),
  ('DAIRY-001-S03','Fence & Corral Repair',
   'The storm damaged a section of fence and a corral gate. You need to replace some boards, reattach the hinge, and close the gap in the wire.',
   'La tormenta dañó una sección de la cerca y la puerta del corral. Tiene que reemplazar unas tablas, volver a fijar la bisagra y cerrar el hueco en el alambre.',
   [
    ('DAIRY-001-Q36','Use the [hammer] to fix the broken gate post.','hammer','martillo',['taladro','desarmador']),
    ('DAIRY-001-Q37','Use the hammer to [fix] the broken gate post.','fix (repair)','arreglar',['limpiar','doblar']),
    ('DAIRY-001-Q38','Use the hammer to fix the broken [gate] post.','gate','puerta',['cerca','establo']),
    ('DAIRY-001-Q39','Use the hammer to fix the broken gate [post].','post','poste',['bisagra','pestillo']),
    ('DAIRY-001-Q40','Use the [screwdriver] to tighten the loose screw on the gate hinge.','screwdriver','desarmador',['martillo','llave']),
    ('DAIRY-001-Q41','Use the screwdriver to [tighten] the loose screw on the gate hinge.','tighten','apretar',['doblar','fijar']),
    ('DAIRY-001-Q42','Use the screwdriver to tighten the loose [screw] on the gate hinge.','screw','tornillo',['clavo','tuerca']),
    ('DAIRY-001-Q43','Use the screwdriver to tighten the loose screw on the gate [hinge].','hinge','bisagra',['pestillo','poste']),
    ('DAIRY-001-Q44','Use the [nail gun] to attach the boards to the fence frame.','nail gun','clavadora',['taladro','martillo']),
    ('DAIRY-001-Q45','Use the nail gun to [attach] the boards to the fence frame.','attach','fijar',['limpiar','apretar']),
    ('DAIRY-001-Q46','Use the nail gun to attach the [boards] to the fence frame.','boards','tablas',['postes','grapas']),
    ('DAIRY-001-Q47','Use the nail gun to attach the boards to the [fence frame].','fence frame','marco de la cerca',['pasillo','poste']),
    ('DAIRY-001-Q48','Use the [staples] to attach the wire to the fence posts.','staples','grapas',['tornillos','cadena']),
    ('DAIRY-001-Q49','Use the staples to attach the [wire] to the fence posts.','wire','alambre',['cadena','cuerda']),
    ('DAIRY-001-Q50','[Tie] the gate with the chain until we get a new latch.','tie','amarrar',['fijar','conectar']),
    ('DAIRY-001-Q51','Tie the gate with the [chain] until we get a new latch.','chain','cadena',['alambre','cuerda']),
    ('DAIRY-001-Q52','Tie the gate with the chain until we get a new [latch].','latch','pestillo',['poste','bisagra']),
    ('DAIRY-001-Q53','Use the wire to [close] the gap in the fence.','close','cerrar',['fijar','amarrar']),
   ]),
  ('DAIRY-001-S04','Field & Equipment',
   'Today you are working in the field. You need to run the chopper, fill the bunker with silage, and do some maintenance checks on the equipment.',
   'Hoy está trabajando en el campo. Tiene que poner en marcha la picadora, llenar la pila con ensilaje y verificar el equipo.',
   [
    ('DAIRY-001-Q54','[Run] the chopper to cut the corn for silage.','run (machine)','poner en marcha',['encender','conectar']),
    ('DAIRY-001-Q55','Run the [chopper] to cut the corn for silage.','chopper','picadora',['cosechadora','mezcladora']),
    ('DAIRY-001-Q56','Run the chopper to [cut] the corn for silage.','cut','cortar',['compactar','limpiar']),
    ('DAIRY-001-Q57','Run the chopper to cut the [corn] for silage.','corn','maíz',['heno','ensilaje']),
    ('DAIRY-001-Q58','Run the chopper to cut the corn for [silage].','silage','ensilaje',['alimento','heno']),
    ('DAIRY-001-Q59','[Drive] the tractor to compact the silage in the bunker.','drive','manejar',['encender','cargar']),
    ('DAIRY-001-Q60','Drive the [tractor] to compact the silage in the bunker.','tractor','tractor',['cargadora','remolque']),
    ('DAIRY-001-Q61','Drive the tractor to [compact] the silage in the bunker.','compact','compactar',['limpiar','cortar']),
    ('DAIRY-001-Q62','Drive the tractor to compact the silage in the [bunker].','bunker','la pila',['establo','laguna']),
    ('DAIRY-001-Q63','Put the [tires] on the tarp over the bunker.','tires','llantas',['pacas','tablas']),
    ('DAIRY-001-Q64','Put the tires on the [tarp] over the bunker.','tarp','la loma',['la laguna','la pila']),
    ('DAIRY-001-Q65','[Check] the grease on the combine axle and the tractor oil.','check','verificar',['arreglar','encender']),
    ('DAIRY-001-Q66','Check the [grease] on the combine axle and the tractor oil.','grease','grasa',['aceite','cinta']),
    ('DAIRY-001-Q67','Check the grease on the [combine] axle and the tractor oil.','combine','cosechadora',['empacadora','mezcladora']),
    ('DAIRY-001-Q68','Check the grease on the combine [axle] and the tractor oil.','axle','eje',['llanta','tuerca']),
    ('DAIRY-001-Q69','Check the grease on the combine axle and the tractor [oil].','oil','aceite',['grasa','agua']),
    ('DAIRY-001-Q70','[Fix] the cable on the reel of the alley scraper.','fix (repair)','arreglar',['encender','verificar']),
    ('DAIRY-001-Q71','Fix the [cable] on the reel of the alley scraper.','cable','cable',['cuerda','manguera']),
    ('DAIRY-001-Q72','Fix the cable on the [reel] of the alley scraper.','reel','carrete',['eje','poste']),
    ('DAIRY-001-Q73','Use the [backhoe] to dig the drainage ditch.','backhoe','excavadora',['minicargadora','tractor']),
    ('DAIRY-001-Q74','Use the backhoe to [dig] the drainage ditch.','dig','cavar',['limpiar','compactar']),
    ('DAIRY-001-Q75','Use the backhoe to dig the drainage [ditch].','ditch','zanja',['pasillo','laguna']),
   ]),
  ('DAIRY-001-S05','Sick Cow / Hospital',
   'A cow is down in her stall. You need to move her to the hospital area, give her medication, and make sure she has feed and water within reach.',
   'Una vaca está caída en su corral. Tiene que moverla al área del hospital, darle medicamento y asegurarse de que tenga alimento y agua a su alcance.',
   [
    ('DAIRY-001-Q76','[Move] the sick cow from her stall to the hospital.','move','mover',['amarrar','limpiar']),
    ('DAIRY-001-Q77','Move the sick cow from her [stall] to the hospital.','stall','corral',['pasillo','establo']),
    ('DAIRY-001-Q78','Move the sick cow from her stall to the [hospital].','hospital','hospital',['laguna','establo']),
    ('DAIRY-001-Q79','Use the [skid steer] to lift and move the cow slowly to the hospital area.','skid steer','minicargadora',['excavadora','cargadora']),
    ('DAIRY-001-Q80','Use the skid steer to [lift] and move the cow slowly to the hospital area.','lift','subir',['mover','lavar']),
    ('DAIRY-001-Q81','Use the skid steer to lift and move the cow [slowly] to the hospital area.','slowly','despacio',['rápido','fuerte']),
    ('DAIRY-001-Q82','[Give] the cow the injection and medication for mastitis.','give','dar',['poner','verificar']),
    ('DAIRY-001-Q83','Give the cow the [injection] and medication for mastitis.','injection','inyección',['medicamento','agua']),
    ('DAIRY-001-Q84','Give the cow the injection and [medication] for mastitis.','medication','medicamento',['inyección','agua']),
    ('DAIRY-001-Q85','Give the cow the injection and medication for [mastitis].','mastitis','mastitis',['el ordeño','la cosecha']),
    ('DAIRY-001-Q86','Give the cow [feed] and water within reach of her head.','feed','alimento',['agua','medicamento']),
    ('DAIRY-001-Q87','Give the cow feed and [water] within reach of her head.','water','agua',['leche','alimento']),
   ]),
]

TOOLS_QUESTIONS = [
    ('TOOLS-001-Q01','Use the [axe] to cut down that small tree.',None,'Usa el hacha para talar ese pequeño árbol.',['Usa la sierra para talar ese pequeño árbol.','Usa el rastrillo para limpiar ese pequeño árbol.','Usa la pala para mover ese pequeño árbol.']),
    ('TOOLS-001-Q02','If the weather is wet, wear the [rubber boots].',None,'Si el clima está húmedo, use botas de goma.',['Si el clima está frío, use guantes de goma.','Si el clima está húmedo, use un impermeable.','Si el clima está seco, use botas de trabajo.']),
    ('TOOLS-001-Q03','Put [fertilizer] on the new plants.',None,'Ponga fertilizante en las nuevas plantas.',['Ponga agua en las nuevas plantas.','Ponga mantillo en las nuevas plantas.','Ponga tierra en las nuevas plantas.']),
    ('TOOLS-001-Q04','The client wants brick [edging] around the large bed.',None,'El cliente quiere un borde de ladrillo alrededor de la cama grande.',['El cliente quiere una cerca de ladrillo alrededor del jardín.','El cliente quiere un camino de ladrillo junto a la cama grande.','El cliente quiere una fuente de ladrillo en la cama grande.']),
    ('TOOLS-001-Q05','Put the tulip bulbs in the [flowerpots].',None,'Ponga los bulbos de tulipán en las macetas.',['Ponga las semillas de tulipán en el suelo.','Saque los bulbos de tulipán de las macetas.','Ponga los bulbos de tulipán en las bolsas.']),
    ('TOOLS-001-Q06','Wear [gloves] to protect your hands.',None,'Use guantes para proteger sus manos.',['Use botas para proteger sus pies.','Use lentes para proteger sus ojos.','Use guantes para limpiar las herramientas.']),
    ('TOOLS-001-Q07','Use the [hoe] to chop those roots.',None,'Use la azada para cortar esas raíces.',['Use el rastrillo para recoger esas raíces.','Use la pala para sacar esas raíces.','Use la azada para regar esas plantas.']),
    ('TOOLS-001-Q08','Bring the long [hoses] and the [hose reel].',None,'Traiga las mangueras largas y el carrete de manguera.',['Traiga los cubos grandes y el aspersor.','Traiga las mangueras largas y el aspersor.','Traiga las regaderas y el carrete de manguera.']),
    ('TOOLS-001-Q09','Bring gas for the [lawn mower].',None,'Traiga gasolina para la cortadora de césped.',['Traiga agua para la cortadora de césped.','Traiga gasolina para el soplador de hojas.','Traiga gasolina para la podadora de setos.']),
    ('TOOLS-001-Q10','Use the [pitchfork] to put the long straw on the tarp.',None,'Use la horca para poner la pajita larga sobre la lona.',['Use la pala para poner la tierra sobre la lona.','Use el rastrillo para poner las hojas sobre la lona.','Use la horca para quitar la pajita de la lona.']),
    ('TOOLS-001-Q11','Use the [pruning saw] for those crossed branches.',None,'Utilice la sierra de podar para esas ramas cruzadas.',['Utilice las tijeras de podar para esas ramas cruzadas.','Utilice la sierra de podar para esas raíces largas.','Utilice el hacha para esas ramas cruzadas.']),
    ('TOOLS-001-Q12','Use the [rake] to get out the old mulch.',None,'Utilice un rastrillo para sacar el mantillo viejo.',['Utilice una pala para sacar el mantillo viejo.','Utilice un rastrillo para poner el mantillo nuevo.','Utilice una horca para sacar el mantillo viejo.']),
    ('TOOLS-001-Q13','Use the [scarifier] to aerate the lawn.',None,'Utilice el escarificador para airear el césped.',['Utilice la cortadora para cortar el césped.','Utilice el escarificador para regar el césped.','Utilice el rastrillo para airear el césped.']),
    ('TOOLS-001-Q14','Use the [pruners] to remove extra branches from the bush.',None,'Utilice las podadoras para quitar las ramas sobrantes del arbusto.',['Utilice la sierra para quitar las ramas sobrantes del arbusto.','Utilice las podadoras para dar forma al arbusto.','Utilice las podadoras para quitar las hojas secas del árbol.']),
    ('TOOLS-001-Q15','Bring the [loppers] and the [electric shears] to shape the hedges.',None,'Traiga las podadoras y las tijeras eléctricas para dar forma a los setos.',['Traiga el rastrillo y la sierra para dar forma a los setos.','Traiga las podadoras y las tijeras eléctricas para cortar el césped.','Traiga las podadoras y la manguera para regar los setos.']),
    ('TOOLS-001-Q16','Use the [shovel] to dig the hole for the root ball.',None,'Utilice la pala para cavar el hoyo para el cepellón.',['Utilice la azada para cavar el hoyo para el cepellón.','Utilice la pala para rellenar el hoyo para el cepellón.','Utilice la pala plana para cavar el hoyo para el cepellón.']),
    ('TOOLS-001-Q17','Use the flat [spade] to edge the bed.',None,'Utilice la pala plana para bordear la cama.',['Utilice la pala para cavar la cama.','Utilice la azada para bordear la cama.','Utilice la pala plana para rellenar la cama.']),
    ('TOOLS-001-Q18','Bring the [sprinkler] to water the new plantings.',None,'Traiga el aspersor para regar las nuevas plantaciones.',['Traiga la regadera para regar las nuevas plantaciones.','Traiga el aspersor para mover las nuevas plantaciones.','Traiga la manguera para regar las nuevas plantaciones.']),
    ('TOOLS-001-Q19','Use the [trowel] to put in the new flowers.',None,'Utilice la paleta para poner las flores nuevas.',['Utilice la pala para poner las flores nuevas.','Utilice la paleta para quitar las flores viejas.','Utilice el rastrillo para poner las flores nuevas.']),
    ('TOOLS-001-Q20','Use the [watering can] for the window boxes.',None,'Utilice la regadera para las jardineras.',['Utilice la manguera para las jardineras.','Utilice la regadera para las macetas grandes.','Utilice el aspersor para las jardineras.']),
    ('TOOLS-001-Q21','Bring the [wheelbarrow] to carry the gravel and sand.',None,'Traiga la carretilla para llevar la grava y la arena.',['Traiga los cubos para llevar la grava y la arena.','Traiga la carretilla para llevar la tierra y el mantillo.','Traiga la carretilla para llevar las plantas y las macetas.']),
]

def bracket_to_u(sentence):
    """Convert [word] authoring syntax to <u>word</u> HTML."""
    import re
    return re.sub(r'\[([^\]]+)\]', r'<u>\1</u>', sentence)

def build_xlsx(path, quiz_id, quiz_title, scenarios=None, flat_questions=None):
    wb = Workbook()
    ws = wb.active
    ws.title = 'quiz_data'
    style_sheet(ws, COLUMNS)
    row = 3

    if scenarios:
        for (sc_id, sc_title, sc_en, sc_es, questions) in scenarios:
            for q in questions:
                q_id, sentence, keyword, correct, distractors = q
                opts, ci = shuffle_options(correct, distractors)
                # Pad to 4 options (some dairy questions only have 2 distractors → 3 options total)
                opt_d = opts[3] if len(opts) > 3 else None
                write_data_row(ws, row, [
                    quiz_id, quiz_title,
                    sc_id, sc_title, sc_en, sc_es,
                    q_id, sentence, keyword,
                    opts[0], opts[1], opts[2], opt_d,
                    'en-US', ''
                ])
                ws.row_dimensions[row].height = 20
                row += 1
    else:
        for q in flat_questions:
            q_id, sentence, keyword, correct, distractors = q
            opts, ci = shuffle_options(correct, distractors)
            opt_d = opts[3] if len(opts) > 3 else None
            write_data_row(ws, row, [
                quiz_id, quiz_title,
                '', '', '', '',
                q_id, sentence, keyword,
                opts[0], opts[1], opts[2], opt_d,
                'en-US', ''
            ])
            ws.row_dimensions[row].height = 20
            row += 1

    wb.save(path)
    print(f'Saved {path}')

def build_dairy_xlsx():
    build_xlsx(
        f'{OUT}/Dairy/dairy_quiz.xlsx',
        'DAIRY-001', 'Dairy Farm Vocabulary Quiz',
        scenarios=DAIRY_SCENARIOS
    )

def build_tools_xlsx():
    build_xlsx(
        f'{OUT}/Landscaping/tools_quiz.xlsx',
        'TOOLS-001', 'Landscaping Tools Quiz',
        flat_questions=TOOLS_QUESTIONS
    )

# ── CSV LIBRARY ──────────────────────────────────────────────────────────────
CSV_FIELDS = [
    'quiz_id','quiz_title','topic_folder',
    'scenario_id','scenario_title',
    'question_id','sentence_en','keyword',
    'correct','option_b','option_c','option_d',
    'tts_lang'
]

def build_csv():
    rows = []
    # Dairy
    for (sc_id, sc_title, _, _, questions) in DAIRY_SCENARIOS:
        for q in questions:
            q_id, sentence, keyword, correct, distractors = q
            opts, _ = shuffle_options(correct, distractors)
            opt_d = opts[3] if len(opts) > 3 else ''
            rows.append({
                'quiz_id': 'DAIRY-001', 'quiz_title': 'Dairy Farm Vocabulary Quiz',
                'topic_folder': 'Dairy',
                'scenario_id': sc_id, 'scenario_title': sc_title,
                'question_id': q_id, 'sentence_en': sentence,
                'keyword': keyword or '', 'correct': correct,
                'option_b': distractors[0] if len(distractors) > 0 else '',
                'option_c': distractors[1] if len(distractors) > 1 else '',
                'option_d': distractors[2] if len(distractors) > 2 else '',
                'tts_lang': 'en-US'
            })
    # Tools
    for q in TOOLS_QUESTIONS:
        q_id, sentence, keyword, correct, distractors = q
        rows.append({
            'quiz_id': 'TOOLS-001', 'quiz_title': 'Landscaping Tools Quiz',
            'topic_folder': 'Landscaping',
            'scenario_id': '', 'scenario_title': '',
            'question_id': q_id, 'sentence_en': sentence,
            'keyword': keyword or '', 'correct': correct,
            'option_b': distractors[0] if len(distractors) > 0 else '',
            'option_c': distractors[1] if len(distractors) > 1 else '',
            'option_d': distractors[2] if len(distractors) > 2 else '',
            'tts_lang': 'en-US'
        })

    with open(f'{OUT}/quiz-library.csv', 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=CSV_FIELDS)
        writer.writeheader()
        writer.writerows(rows)
    print(f'CSV library: {len(rows)} questions written.')

# ── HTML GENERATION ──────────────────────────────────────────────────────────
# Regenerate both HTML files from the authoritative data above,
# with options baked in their shuffled-at-generation order.

def build_dairy_html():
    """Rebuild dairy_quiz.html with static baked option order."""
    scenarios_js = []
    random.seed(42)
    for (sc_id, sc_title, sc_en, sc_es, questions) in DAIRY_SCENARIOS:
        narrative = f'{sc_en}<br><em>{sc_es}</em>'
        qs_js = []
        for q in questions:
            q_id, sentence, keyword, correct, distractors = q
            opts, ci = shuffle_options(correct, distractors)
            sentence_html = bracket_to_u(sentence)
            opt_strs = ',\n        '.join(f"'{o}'" for o in opts)
            qs_js.append(f"""      {{ sentence: '{sentence_html}',
        keyword: '{keyword}', options: [{opt_strs}], correctIndex: {ci} }}""")
        scenarios_js.append(f"""  {{
    title: "{sc_title}",
    narrative: "{narrative}",
    questions: [
{chr(10).join(qs_js)}
    ]
  }}""")
    scenarios_block = 'const scenarios = [\n' + ',\n'.join(scenarios_js) + '\n]; // end scenarios'
    return scenarios_block

def build_tools_html():
    """Rebuild tools_quiz.html with static baked option order."""
    random.seed(42)
    qs_js = []
    for q in TOOLS_QUESTIONS:
        q_id, sentence, keyword, correct, distractors = q
        opts, ci = shuffle_options(correct, distractors)
        sentence_html = bracket_to_u(sentence)
        opt_strs = ',\n    '.join(f"'{o}'" for o in opts)
        qs_js.append(f"""  {{ eng: '{sentence_html}',
    options: [{opt_strs}], correctIndex: {ci} }}""")
    return 'const questions = [\n' + ',\n'.join(qs_js) + '\n]; // end questions'

if __name__ == '__main__':
    build_template()
    build_dairy_xlsx()
    build_tools_xlsx()
    build_csv()
    print('\nData blocks for HTML:')
    print('--- DAIRY ---')
    print(build_dairy_html()[:300], '...')
    print('--- TOOLS ---')
    print(build_tools_html()[:300], '...')
    print('\nDone.')
